""" Module to retrieve financial data of hospitals """

# External imports
import openpyxl as pyxl

# Internal imports
from common import config, helpers

# Internal functions
def read_financial_data(actual_year, data_frame, hospital_ids):
    """ Reads xlsm file to retrieve
    the financial data of the hospitals
    """
    print("a")
    workbook = pyxl.load_workbook(f'../../raw_data/financial_data/{actual_year}.xlsm',
                                  data_only= True, read_only=True)
    print("b")
    sheet = workbook[config.FINANCIAL_DATA_SHEET]
    print("c")
    results_row = get_accumulated_results_row(sheet, actual_year, data_frame)
    print("d")
    target_row = results_row + 2 # 21 and 22 detail is in the next row
    print("e")
    index_21, index_22 = get_data_columns_index(sheet, target_row)
    print("f")
    data_frame['21_value'] = '--'
    print("g")
    data_frame['22_value'] = '--'
    print("h")
    for index in hospital_ids:
        print(index)
        if index == '--':
            continue

        try:
            data_frame.loc[index, '21_value'] = round(sheet.cell(row=data_frame.loc[index, 'row_index'],
                                                                                        column=index_21).value)
            #print("0")
            data_frame.loc[index, '22_value'] = round(sheet.cell(row=data_frame.loc[index, 'row_index'],
                                                                                        column=index_22).value)
        except Exception:
            print("ERROR EN: ",index)
            continue

        #print("2")
def get_accumulated_results_row(sheet, actual_year, data_frame):
    """ Gets the row index where the accumulated
    information starts
    """

    accumulated_results_row = -2
    hospital_fns_list = data_frame.index.tolist()
    data_frame['row_index'] = '--'
    hospital_fns_idexes = []

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        if len(hospital_fns_list) == len(hospital_fns_idexes):
            break
        for cell in row:
            if cell.value == f'ACUMULADO  {actual_year}':
                accumulated_results_row = cell.row
            if (cell.value in hospital_fns_list) and (accumulated_results_row > 0) :
                if cell.value not in hospital_fns_idexes:
                    hospital_fns_idexes.append(cell.value)
                    data_frame.loc[cell.value, 'row_index'] = cell.row
            if len(hospital_fns_list) == len(hospital_fns_idexes):
                break

    return accumulated_results_row

def get_data_columns_index(sheet, target_row):
    """ Gets the column index where the accumulated
    information of detail 21 and 22 are
    """
    index_21 = 0
    index_22 = 0
    for col_idx, cell in enumerate(sheet[target_row], start=1):
        if cell.value == 21:
            index_21 = col_idx
        if cell.value == 22:
            index_22 = col_idx
    return index_21, index_22

# Body

hospital_dataframe = helpers.dataframe_read_csv_file(path=config.HOSPITAL_DICTIONARY_PATH)
years = config.YEARS_RANGE

print("\033[92mRetrieving financial data...\033[0m")

for year in years:
    print("1")
    helpers.dataframe_rewrite_index(data_frame=hospital_dataframe, new_index='hospital_fns_id')
    print("2")
    fns_ids = helpers.dataframe_index_to_list(data_frame=hospital_dataframe)
    print("3")
    read_financial_data(year, hospital_dataframe, fns_ids)
    print("4")
    helpers.dataframe_reset_index(data_frame=hospital_dataframe)
    print("5")
    selected_columns = ['hospital_id', 'hospital_fns_id', 'region',
                        'row_index', '21_value', '22_value']
    print("6")
    column_filtered_df = helpers.dataframe_filter_columns(data_frame=hospital_dataframe,
                                                          columns=selected_columns)
    print("7")
    helpers.dataframe_write_csv_file(data_frame=column_filtered_df,
                                     path=f'../../data/input/{year}.csv')

    print(f"\t\033[93mYear: {year} done!\033[93m")

print("\033[92mFinancial data retrieved!\033[0m")
